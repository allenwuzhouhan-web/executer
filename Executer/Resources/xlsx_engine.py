#!/usr/bin/env python3
"""
Executer XLSX Engine
Creates .xlsx files from JSON spec with styling, formulas, and auto-sizing.

Usage:
    python3 xlsx_engine.py --spec spec.json --output output.xlsx
    echo '{"sheets":[...]}' | python3 xlsx_engine.py --stdin --output output.xlsx
"""
import argparse, importlib.util, json, sys, tempfile
from pathlib import Path

# Import image_utils from same directory
try:
    from image_utils import resolve_image, cleanup_temp_images
except ImportError:
    _iu_path = Path(__file__).parent / "image_utils.py"
    if _iu_path.exists():
        _spec = importlib.util.spec_from_file_location("image_utils", _iu_path)
        _mod = importlib.util.module_from_spec(_spec)
        _spec.loader.exec_module(_mod)
        resolve_image = _mod.resolve_image
        cleanup_temp_images = _mod.cleanup_temp_images
    else:
        resolve_image = lambda source, temp_dir=None: source if source and Path(source).exists() else None
        cleanup_temp_images = lambda d: None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    _HAS_XLIMAGE = True
except ImportError as _ie:
    if "openpyxl" in str(_ie):
        print(json.dumps({"success": False, "error": "openpyxl not installed. Run: pip3 install openpyxl"}))
        sys.exit(1)
    _HAS_XLIMAGE = False

def hex_to_argb(hex_str):
    h = hex_str.lstrip("#")
    if len(h) != 6: return "FF000000"
    return "FF" + h.upper()

def log(msg):
    print(msg, file=sys.stderr)

def create_xlsx(spec, output_path):
    try:
        _img_temp_dir = tempfile.mkdtemp(prefix="executer_xlsx_imgs_")
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        sheets = spec.get("sheets", [])
        if not sheets:
            return {"success": False, "error": "No sheets in spec"}

        for sheet_spec in sheets:
            name = sheet_spec.get("name", "Sheet1")
            ws = wb.create_sheet(title=name[:31])

            headers = sheet_spec.get("headers", [])
            rows = sheet_spec.get("rows", [])
            col_widths = sheet_spec.get("column_widths", [])
            header_style = sheet_spec.get("header_style", {})

            header_font = Font(
                bold=header_style.get("bold", True),
                color=hex_to_argb(header_style.get("font_color", "FFFFFF")),
                name=header_style.get("font", "Helvetica Neue"),
                size=header_style.get("font_size", 11),
            )
            header_fill = PatternFill(
                start_color=hex_to_argb(header_style.get("bg_color", "0066CC")),
                end_color=hex_to_argb(header_style.get("bg_color", "0066CC")),
                fill_type="solid",
            )
            alt_fill = PatternFill(start_color="FFF5F5F5", end_color="FFF5F5F5", fill_type="solid")
            thin_border = Border(
                left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"),
                top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD"),
            )
            body_font = Font(name="Helvetica Neue", size=11)

            row_idx = 1
            if headers:
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=str(header))
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                row_idx = 2

            for data_row_idx, row_data in enumerate(rows):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if isinstance(value, (int, float)):
                        cell.value = value
                    elif isinstance(value, str) and value.startswith("="):
                        cell.value = value
                    elif value is None:
                        cell.value = ""
                    else:
                        cell.value = str(value)
                    cell.font = body_font
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")
                    if data_row_idx % 2 == 1:
                        cell.fill = alt_fill
                row_idx += 1

            num_cols = max(len(headers), max((len(r) for r in rows), default=0)) if (rows or headers) else 0
            for col_idx in range(1, num_cols + 1):
                col_letter = get_column_letter(col_idx)
                if col_idx - 1 < len(col_widths):
                    ws.column_dimensions[col_letter].width = col_widths[col_idx - 1]
                else:
                    max_len = 8
                    if col_idx <= len(headers):
                        max_len = max(max_len, len(str(headers[col_idx - 1])) + 2)
                    for rd in rows:
                        if col_idx <= len(rd):
                            max_len = max(max_len, len(str(rd[col_idx - 1])) + 2)
                    ws.column_dimensions[col_letter].width = min(max_len, 50)

            if headers:
                ws.freeze_panes = "A2"

            # Images
            images = sheet_spec.get("images", [])
            if images and _HAS_XLIMAGE:
                for img_spec in images:
                    if isinstance(img_spec, str):
                        img_spec = {"url": img_spec, "cell": "A1"}
                    img_source = img_spec.get("url") or img_spec.get("path", "")
                    local_path = resolve_image(img_source, _img_temp_dir)
                    if local_path and Path(local_path).exists():
                        try:
                            xl_img = XLImage(local_path)
                            if img_spec.get("width"):
                                xl_img.width = img_spec["width"]
                            if img_spec.get("height"):
                                xl_img.height = img_spec["height"]
                            ws.add_image(xl_img, img_spec.get("cell", "A1"))
                        except Exception as e:
                            log(f"Warning: Could not add image: {e}")

            log(f"Sheet '{name}': {len(headers)} cols, {len(rows)} rows, {len(images)} images")

        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(out))
        cleanup_temp_images(_img_temp_dir)
        log(f"Saved: {out}")
        return {"success": True, "path": str(out.resolve()), "sheets": len(sheets)}
    except Exception as e:
        return {"success": False, "error": f"{type(e).__name__}: {str(e)}"}

def main():
    parser = argparse.ArgumentParser(description="Executer XLSX Engine")
    parser.add_argument("--spec", help="Path to spec JSON file")
    parser.add_argument("--stdin", action="store_true", help="Read spec from stdin")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    args = parser.parse_args()

    if args.stdin:
        try: spec = json.load(sys.stdin)
        except json.JSONDecodeError as e:
            print(json.dumps({"success": False, "error": f"Invalid JSON: {e}"})); sys.exit(1)
    elif args.spec:
        try: spec = json.loads(Path(args.spec).read_text(encoding="utf-8"))
        except Exception as e:
            print(json.dumps({"success": False, "error": str(e)})); sys.exit(1)
    else:
        print(json.dumps({"success": False, "error": "Provide --spec or --stdin"})); sys.exit(1)

    output = args.output
    if output.endswith("/") or Path(output).is_dir():
        filename = spec.get("filename", "spreadsheet.xlsx")
        if not filename.endswith(".xlsx"): filename += ".xlsx"
        output = str(Path(output) / filename)

    result = create_xlsx(spec, output)
    print(json.dumps(result))
    sys.exit(0 if result["success"] else 1)

if __name__ == "__main__":
    main()
